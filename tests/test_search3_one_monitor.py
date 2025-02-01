import json
from student_tree_search import *
from colorama import Fore, Style
import inspect
import math
import pandas as pd
from time import monotonic
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import concurrent.futures
import signal
import sys
import threading
import multiprocessing
import psutil
import os

def run_searchProcess(problem, search_strategy, search_limit,search_improve, search_heuristic, queue):
    """Método executado no processo filho."""
    try:
        print(f"Search: {search_strategy}, {search_limit}, {search_improve}, {search_heuristic} -- iniciada")
        searchTree = SearchTree(problem, search_strategy, search_improve, search_limit, search_heuristic)
        path = searchTree.search()
        stats = searchTree.get_stats()
    except Exception as e:
        print(f"Erro durante a busca: {e}")
        path = None
        stats = None
    queue.put((path, stats))
def run_monitorProcess(monitored_pids, max_cpu_time=5.0, monitorization_interval=0.5, total_cpu_time_queue=None):
        """
        Monitora uso de CPU do processo (PID=monitored_pid). 
        Se (user + system) >= max_cpu_time, o processo é finalizado (kill).
        """
        ps_processes = []
        for pid in monitored_pids:
            try:
                p = psutil.Process(pid)
                ps_processes.append(p)
            except psutil.NoSuchProcess:
                pass  # o processo já terminou antes de começarmos a monitorizar

        while True:
            for monitored_process in ps_processes:
                try:
                    cpu_times = monitored_process.cpu_times()
                    user_time = cpu_times.user
                    system_time = cpu_times.system
                    total_cpu_time = user_time + system_time
                    print(
                        f"[Monitor] user={user_time:.2f}s | system={system_time:.2f}s | total={total_cpu_time:.2f}s"
                    )
                    if total_cpu_time >= max_cpu_time:
                        print(f"Tempo de CPU chegou ao limite de {max_cpu_time}s. Encerrando processo-alvo.")
                        total_cpu_time_queue.put([monitored_process.pid,total_cpu_time])
                        monitored_process.kill()  # ou p.terminate() se preferir
                        break
                except psutil.NoSuchProcess:
                    # O processo alvo terminou antes de atingir o limite
                    print("[Monitor] Processo alvo finalizou, encerrando monitor.")
                except KeyboardInterrupt:
                    pass
            time.sleep(monitorization_interval)
class MultiSearchProcessesManager:
    """ def __init__(self, problem, search_strategy, search_limit, search_improve, search_heuristic, scene):
        self.problem = problem
        self.search_strategy = search_strategy
        self.search_limit = search_limit
        self.search_improve = search_improve
        self.search_heuristic = search_heuristic
        self.path = None
        self.stats = None
        self.search_time = 0
        # atributos para launch()
        self.result_queue = None
        self.total_cpu_time_value = None
        self.search_process = None
        self.monitor_process = None
        self.scene = scene """
    def __init__(self):
        self.search_jobs = []  # lista de dicts com info de cada busca
        self.processes = []    # lista de Process
        self.queues = []       # cada job terá a sua Queue ([pid,result_queue]) precisa do pid para depois poder associar a queue dos cpu_times do monitor
        self.monitor_process = None
        self.total_cpu_time_queue = multiprocessing.Queue()
        self.all_results = []

    def add_searchProcess(self, problem, search_strategy, search_limit, search_improve, search_heuristic, scene_name):
        job = {
            'problem': problem,
            'strategy': search_strategy,
            'limit': search_limit,
            'improve': search_improve,
            'heuristic': search_heuristic,
            'scene': scene_name
        }
        self.search_jobs.append(job)
    def start_searchProcesses(self,max_cpu_time=5.0, monitorization_interval=0.5):
        for job in self.search_jobs:
            result_queue = multiprocessing.Queue()
            p = multiprocessing.Process(
                target=run_searchProcess,
                args=(
                    job['problem'],
                    job['search_strategy'],
                    job['search_limit'],
                    job['search_improve'],
                    job['search_heuristic'],
                    result_queue
                )
            )
            p.start()
            self.processes.append(p)
            self.queues.append([p.pid,result_queue])
        
        pids = [p.pid for p in self.processes if p.is_alive()] # ficar com os pids para o monitor
        self.monitor_process = multiprocessing.Process(
            target=run_monitorProcess,
            args=(pids, max_cpu_time, monitorization_interval, self.total_cpu_time_queue)
        )
        self.monitor_process.start()
        
    def wait(self):
        # Esperar que cada processo de busca termine
        for p in self.processes:
            p.join()
            exit_code = p.exitcode
            print(f" terminou com exitcode={exit_code}")
        
        # Finaliza o monitor
        if self.monitor_process.is_alive():
            self.monitor_process.terminate()
        self.monitor_process.join()

        all_results = []
        for pid_queue, job in zip(self.queues, self.search_jobs):
            pid, queue = pid_queue
            p_cpu_time = None
            for cpu_time in self.total_cpu_time_queue:
                if cpu_time[0] == pid:
                    p_cpu_time = cpu_time[1]
            if not queue.empty():                
                path, stats = queue.get()
                all_results.append({
                    'scene': job['scene'],
                    'strategy': job['search_strategy'],
                    'limit': job['search_limit'],
                    'improve': job['search_improve'],
                    'heuristic': job['search_heuristic'],
                    'path': path,
                    'stats': stats,
                    'search_time': p_cpu_time
                })
            else:
                all_results.append({
                    'scene': job['scene'],
                    'strategy': job['search_strategy'],
                    'limit': job['search_limit'],
                    'improve': job['search_improve'],
                    'heuristic': job['search_heuristic'],
                    'path': None,
                    'stats': None,
                    'search_time': p_cpu_time
                })
        print("Fim do processo de busca e monitor.")
        self.all_results = all_results
def append_df_to_excel(file_name, df, scene_name):
    """
    Esta função escreve a 'scene_name' numa linha,
    e depois escreve o DF logo abaixo.
    """
    # 1) Se o ficheiro Excel não existir, cria um Excel vazio
    if not os.path.exists(file_name):
        df_vazio = pd.DataFrame()
        df_vazio.to_excel(file_name, index=False)
    
    # 2) Abrir o Excel com openpyxl
    wb = load_workbook(file_name)
    ws = wb.active # ou wb[sheet_name] se quiser uma folha específica
    
    # 3) Encontrar última linha ocupada
    last_row = ws.max_row
    
    # 1) Escrever a linha com o nome da scene na 1ª coluna
    ws.cell(row=last_row+1, column=1, value=f"Scene: {scene_name}")
    
    # 5) Gravar (anexar) o DataFrame abaixo dessa linha.
    #    Para isso, usamos pd.ExcelWriter em modo "append" (mode='a') e if_sheet_exists="overlay".
    #    Note que NÃO fazemos writer.book = wb.
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # Usamos sheet_name=ws.title para escrever na MESMA folha que já estamos a usar.
        # E definimos startrow = last_row + 2 para que o cabeçalho do DF fique na linha imediatamente abaixo do "Scene: ..."
        df.to_excel(writer,
                    sheet_name=ws.title,
                    startrow=last_row + 2,
                    index=False,
                    header=True)
    
    # Passo 5: Reabrir para aplicar coloração de célula
    wb = load_workbook(file_name)
    ws = wb.active

    if 'Solution' not in df.columns:
        # Se não existir 'Solution', não precisamos colorir nada
        wb.save(file_name)
        return

    # 7) Encontrar o índice da coluna "Solution"
    col_solution_index = list(df.columns).index('Solution') + 1  # +1 pois openpyxl usa índices baseados em 1

    fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="lightGrid")

    # 8) Saber quantas linhas o DF tem
    rows_count = len(df)

    # A linha com "Scene: ..." foi last_row + 1
    # O cabeçalho do DF começa em last_row + 2
    # As linhas de dados do DF vão do (last_row + 3) até (last_row + 2 + rows_count)
    inicio_df = last_row + 1      # primeira linha de dados (abaixo do header)
    fim_df = last_row + 2 + rows_count

    # 9) Iterar pelas linhas recém escritas
    for row in ws.iter_rows(min_row=inicio_df, max_row=fim_df,
                            min_col=1, max_col=ws.max_column):
        valor_solution = row[col_solution_index - 1].value
        if valor_solution == 'FAIL' or valor_solution == 'TIMEOUT':
            for cell in row:
                cell.fill = fill

    wb.save(file_name)
#+: faz todos os processos ao mesmo tempo
def test_search_queue(json_states = 'test_states.json', scenes = ["scene_0"], excel_name = "resultados.xlsx",max_cpu_time=5.0, monitorization_interval=0.5):
    search_strategies = ['greedy','informed','a*']#['informed','a*','greedy','uniform','depth','breadth']
    search_improves = [False, True]
    search_limits = [10, 50, None]
    search_heuristics = ['manhattan']
    
    manager = MultiSearchProcessesManager()
    try:
        with open(json_states, 'r') as file:
            all_scenes = json.load(file)
    except Exception as e:
        print(f"Erro ao carregar o JSON: {e}")
        return
    for scene in scenes:  
        if "_goal" in scene:
            continue
        goal_key = f"{scene}_goal"
        if goal_key not in all_scenes:
            print(f"Não foi encontrada a key '{goal_key}' para a scene '{scene}'.")
            continue
        states = all_scenes[scene]
        goal = all_scenes[goal_key]
        domain = SearchDomain()
        for chave in states:
            state = states[chave]
            domain.atualize_domain(state)
            print(domain.body)
        problem = SearchProblem(domain, goal)
        print(problem.initial)
        print(problem.domain.head)
        resultados_scene = []
        processos = []  # para guardar todos os SearchProcess
        for search_strategy in search_strategies:
            for search_limit in search_limits:
                for search_improve in search_improves:
                    for search_heuristic in search_heuristics:
                        manager.add_searchProcess(problem, search_strategy, search_limit, search_improve, search_heuristic, scene)
             
        manager.start_all(max_cpu_time=5.0, monitor_interval=0.5)
        manager.wait_all() 
        # 5) Depois de gerar os resultados da scene, convertemos num DataFrame
        df_scene = pd.DataFrame(manager.all_results )
        # 6) Guardamos (anexamos) no Excel, chamando a função auxiliar
        append_df_to_excel(excel_name, df_scene, scene)
        #domain.print_mapa(path = path)
    print("Fim do teste.")
def test_reta_heuristic():
    def rotacionar_ponto(ponto, angulo, ponto_medio):
        # Transladar ponto para a origem
        x, y = ponto
        px, py = ponto_medio
        x_transladado = x - px
        y_transladado = y - py

        # Aplicar matriz de rotação
        angulo_rad = math.radians(angulo)
        x_rotacionado = x_transladado * math.cos(angulo_rad) - y_transladado * math.sin(angulo_rad)
        y_rotacionado = x_transladado * math.sin(angulo_rad) + y_transladado * math.cos(angulo_rad)

        # Transladar ponto de volta
        x_final = x_rotacionado + px
        y_final = y_rotacionado + py

        return (x_final, y_final)
    def rotacionar_reta(ponto1, ponto2, angulo):
        # Calcular ponto médio
        ponto_medio = ((ponto1[0] + ponto2[0]) / 2, (ponto1[1] + ponto2[1]) / 2)

        # Rotacionar pontos
        ponto1_rotacionado = rotacionar_ponto(ponto1, angulo, ponto_medio)
        ponto2_rotacionado = rotacionar_ponto(ponto2, angulo, ponto_medio)

        return ponto1_rotacionado, ponto2_rotacionado
    def calcular_reta(ponto1, ponto2):
        x1, y1 = ponto1
        x2, y2 = ponto2

        # Calcula a inclinação (m)
        if x2 - x1 != 0:
            m = (y2 - y1) / (x2 - x1)
        else:
            raise ValueError("Os pontos têm a mesma coordenada x, a inclinação é indefinida (reta vertical).")

        # Calcula a interseção com o eixo y (b)
        b = y1 - m * x1

        return m, b

    def pontos_inteiros_na_reta(ponto1, ponto2):
        m, b = calcular_reta(ponto1, ponto2)
        print(f" y = {m}x + {b}")
        x1, y1 = ponto1
        x2, y2 = ponto2

        pontos_inteiros = []

        # Iterar sobre os valores de x entre x1 e x2
        for x in range(min(x1, x2), max(x1, x2) + 1):
            y = m * x + b
            ponto = [x, int(y)]
            if ponto not in pontos_inteiros:
                pontos_inteiros.append(ponto)

        # Iterar sobre os valores de y entre y1 e y2
        for y in range(min(y1, y2), max(y1, y2) + 1):
            if m != 0:
                x = (y - b) / m
                ponto = [int(x), y]
                if ponto not in pontos_inteiros:
                    pontos_inteiros.append(ponto)

        pontos_inteiros.sort()
        return pontos_inteiros
    domain = SearchDomain()
    #state = {"size": [10, 10], "map":[[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0]], "fps": 10, "timeout": 3000, "level": 1}
    state = {"size": [20, 20], "map":[[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]], "fps": 10, "timeout": 3000, "level": 1}
    domain.atualize_domain(state)
    # Exemplo de uso
    pontos = []
    head = [5, 5]
    #goal = [20, 20]
    #goal = [20,10]
    goal = [20,5]
    pontos_goal = pontos_inteiros_na_reta(head, goal)
    print(f"Pontos inteiros na reta entre {head} e {goal}: {pontos_goal}")
    goal_rotated1 = rotacionar_ponto(goal, -40, head)
    print("goal_rotated1",goal_rotated1)
    pontos_goal1 = pontos_inteiros_na_reta(head, (int(round(goal_rotated1[0])),int(round(goal_rotated1[1]))))
    goal_rotated2 = rotacionar_ponto(goal, -20, head)
    print("goal_rotated1",goal_rotated2)
    pontos_goal2 = pontos_inteiros_na_reta(head, (int(round(goal_rotated2[0])),int(round(goal_rotated2[1]))))
    goal_rotated3 = rotacionar_ponto(goal, 20, head)
    print("goal_rotated1",goal_rotated3)
    pontos_goal3 = pontos_inteiros_na_reta(head, (int(round(goal_rotated3[0])),int(round(goal_rotated3[1]))))
    goal_rotated4 = rotacionar_ponto(goal, 40, head)
    print("goal_rotated1",goal_rotated4)
    pontos_goal4 = pontos_inteiros_na_reta(head, (int(round(goal_rotated4[0])),int(round(goal_rotated4[1]))))
    pontos = pontos + pontos_goal
    pontos = pontos + pontos_goal1
    pontos = pontos + pontos_goal2
    pontos = pontos + pontos_goal3
    pontos = pontos + pontos_goal4
    domain.print_mapa(path = pontos)
    
if __name__ == "__main__":
    #ToDo: juntar todos os processos de todas as cenas e usar uma pool de processos para garantir que nao se abrem demasiados processos
    """ start_time = monotonic()
    test_search()
    start_end1_time = monotonic() """
    #ToDo: fazer por scene no excel
    test_search_queue(scenes = ["scene_0","scene_1"], excel_name="resultados2.xlsx")
    """ end_time = monotonic()
    print(f"test_search finalizado em: {start_end1_time-start_time:.2f} sec")
    print(f"test_search_queue finalizado em: {end_time-start_end1_time:.2f} sec") """
    #test_reta()
    #test_search(excel_name="resultados.xlsx")