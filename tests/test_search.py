import json
from student_tree_search import *
from colorama import Fore, Style
import inspect
import math
import pandas as pd
from time import monotonic
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import concurrent.futures
import signal
import sys
import threading
import multiprocessing
import psutil

def run_searchProcess(problem, search_strategy, search_limit,search_improve, search_heuristic, queue):
    """Método executado no processo filho."""
    try:
        print(f"Search: {search_strategy}, {search_limit}, {search_improve}, {search_heuristic} -- iniciada")
        searchTree = SearchTree(problem, search_strategy, search_improve, search_limit, search_heuristic)
        path = searchTree.search()
        stats = searchTree.get_stats()
    except Exception as e:
        print(f"Erro durante a busca: {e}")
        path = None
        stats = None
    queue.put((path, stats))
def run_monitorProcess(monitored_pid, max_cpu_time=5.0, monitorization_interval=0.5, total_cpu_time_value=None):
        """
        Monitora uso de CPU do processo (PID=monitored_pid). 
        Se (user + system) >= max_cpu_time, o processo é finalizado (kill).
        """
        try: # para caso o processo alvo tenha terminado antes
            monitored_process = psutil.Process(monitored_pid)
        except psutil.NoSuchProcess:
            print(f"[Monitor] O processo PID={monitored_pid} não existe (pode ter finalizado antes). Encerrando monitor.")
            return  # sai imediatamente
        try:
            while True:
                cpu_times = monitored_process.cpu_times()
                user_time = cpu_times.user
                system_time = cpu_times.system
                total_cpu_time = user_time + system_time
                print(
                    f"[Monitor] user={user_time:.2f}s | system={system_time:.2f}s | total={total_cpu_time:.2f}s"
                )
                total_cpu_time_value.value = total_cpu_time
                if total_cpu_time >= max_cpu_time:
                    print(f"Tempo de CPU chegou ao limite de {max_cpu_time}s. Encerrando processo-alvo.")
                    monitored_process.kill()  # ou p.terminate() se preferir
                    break
                time.sleep(monitorization_interval)
        except psutil.NoSuchProcess:
            # O processo alvo terminou antes de atingir o limite
            print("[Monitor] Processo alvo finalizou, encerrando monitor.")
        except KeyboardInterrupt:
            pass
class SearchProcess:
    def __init__(self, problem, search_strategy, search_limit, search_improve, search_heuristic):
        self.problem = problem
        self.search_strategy = search_strategy
        self.search_limit = search_limit
        self.search_improve = search_improve
        self.search_heuristic = search_heuristic
        self.path = None
        self.stats = None
        self.search_time = 0
        # atributos para launch()
        self.result_queue = None
        self.total_cpu_time_value = None
        self.search_process = None
        self.monitor_process = None

    def start_searchProcess(self,max_cpu_time=5.0, monitorization_interval=0.5):
        """Inicia um processo de busca e espera pelo seu término."""
        result_queue = multiprocessing.Queue()
        total_cpu_time_value = multiprocessing.Value('d', 0.0)
        search_process = multiprocessing.Process(target=run_searchProcess, args=(self.problem, self.search_strategy, self.search_limit, self.search_improve, self.search_heuristic, result_queue))
        search_process.start()
        # Cria e inicia o monitor
        monitor_process = multiprocessing.Process(target=run_monitorProcess, args=(search_process.pid, max_cpu_time, monitorization_interval, total_cpu_time_value))
        monitor_process.start()
        # Aguarda o término do processo de trabalho (quando for morto pelo monitor, ou terminar normalmente)
        search_process.join()
        exit_code = search_process.exitcode
        print(f"search_process terminou com exitcode={exit_code}")
        if exit_code == 0:
            # Em princípio, o processo terminou sem ser morto.
            # Podemos ler da fila sem bloquear (pois esperamos que algo tenha sido colocado)
            if not result_queue.empty():
                self.path, self.stats = result_queue.get()
                #print(f"search_process retornou path={self.path}, stats={self.stats}")
            else:
                # Se por algum motivo a fila estiver vazia,
                self.path, self.stats = None, None
        else:
            # Sinaliza que foi morto (SIGKILL / erro) ou terminou com erro
            self.path, self.stats = None, None
            
        # Depois que o trabalho termina, finalizamos o monitor (se ainda estiver vivo)
        if monitor_process.is_alive():
            monitor_process.terminate()
        monitor_process.join()
        self.search_time = total_cpu_time_value.value
        print("Fim do programa principal.")
        
    def launch(self,max_cpu_time=5.0, monitorization_interval=0.5):
        """Apenas inicia (start) os processos, sem aguardar finalização."""
        self.result_queue = multiprocessing.Queue()
        self.total_cpu_time_value = multiprocessing.Value('d', 0.0)

        self.search_process = multiprocessing.Process(
            target=run_searchProcess,
            args=(
                self.problem, 
                self.search_strategy, 
                self.search_limit, 
                self.search_improve, 
                self.search_heuristic, 
                self.result_queue
            )
        )
        self.search_process.start()

        self.monitor_process = multiprocessing.Process(
            target=run_monitorProcess,
            args=(
                self.search_process.pid,
                max_cpu_time,
                monitorization_interval,
                self.total_cpu_time_value
            )
        )
        self.monitor_process.start()

    def wait(self):
        """Espera (join) os processos, pega resultados e finaliza o monitor."""
        # Espera o término do processo de busca
        self.search_process.join()
        exit_code = self.search_process.exitcode
        print(f"search_process terminou com exitcode={exit_code}")

        if exit_code == 0:
            # terminou sem ser morto
            if not self.result_queue.empty():
                self.path, self.stats = self.result_queue.get()
            else:
                self.path, self.stats = None, None
        else:
            # morto ou erro
            self.path, self.stats = None, None
        
        # Finaliza o monitor
        if self.monitor_process.is_alive():
            self.monitor_process.terminate()
        self.monitor_process.join()

        # Copia o tempo de CPU medido
        self.search_time = self.total_cpu_time_value.value

        print("Fim do processo de busca e monitor.")

        
def test_search(scene = "scene_0", exel_name = "resultados.xlsx",max_cpu_time=5.0, monitorization_interval=0.5):
    search_strategies = ['greedy','informed','a*']#['informed','a*','greedy','uniform','depth','breadth']
    search_improves = [False, True]
    search_limits = [10, 50, None]
    search_heuristics = ['manhattan']
    resultados = []
    
    domain = SearchDomain()
    try:
        with open('test_states.json', 'r') as file:
            scenes = json.load(file)
            states = scenes[scene]
    except Exception as e:
        print(f"Erro ao carregar o JSON: {e}")
    for chave in states:
        state = states[chave]
        domain.atualize_domain(state)
        print(domain.body)
    problem = SearchProblem(domain, [[23,11]])
    print(problem.initial)
    print(problem.domain.head)
    for search_strategy in search_strategies:
        for search_limit in search_limits:
            for search_improve in search_improves:
                for search_heuristic in search_heuristics:
                    print(f"Search: {search_strategy}, {search_limit}, {search_improve}, {search_heuristic}")
                    search_process = SearchProcess(problem, search_strategy, search_limit, search_improve, search_heuristic)
                    search_process.start_searchProcess(max_cpu_time=5.0, monitorization_interval=0.5)
                    if search_process.stats is not None:
                        stats = search_process.stats
                        stats['elapsed_time'] = search_process.search_time
                    else:
                        stats = {
                            'Strategy': search_strategy,
                            'Improve': search_improve,
                            'Limit': search_limit,
                            'Heuristic': search_heuristic,
                            'Solution': 'TIMEOUT',
                            'Terminals': 1,
                            'Non-terminals': 'None',
                            'Avg Branching': 'None',
                            'Cost': 'None',
                            'Length': 'None',
                            'Num Skipped':  0,
                            'Num Solutions': 'None',
                            'elapsed_time': search_process.search_time,
                            'Path': 'None',
                        }
                    print(stats)
                    resultados.append(stats)
                    # Escrever resultados no Excel
                    df = pd.DataFrame(resultados)
                    df.to_excel(exel_name, index=False)
                    wb = load_workbook(exel_name)
                    ws = wb.active
                    # Aplicar cor de preenchimento às células onde 'Solution' é 'ERROR: None'
                    fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="lightGrid")
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        if row[4].value == 'FAIL' or row[4].value == 'TIMEOUT':  # Supondo que 'Solution' está na segunda coluna
                            for cell in row:
                                cell.fill = fill
                    wb.save(exel_name)
    #domain.print_mapa(path = path)
def test_search_queue(scene = "scene_0", exel_name = "resultados.xlsx",max_cpu_time=5.0, monitorization_interval=0.5):
    search_strategies = ['greedy','informed','a*']#['informed','a*','greedy','uniform','depth','breadth']
    search_improves = [False, True]
    search_limits = [10, 50, None]
    search_heuristics = ['manhattan']
    resultados = []
    
    domain = SearchDomain()
    try:
        with open('test_states.json', 'r') as file:
            scenes = json.load(file)
            states = scenes[scene]
    except Exception as e:
        print(f"Erro ao carregar o JSON: {e}")
    for chave in states:
        state = states[chave]
        domain.atualize_domain(state)
        print(domain.body)
    problem = SearchProblem(domain, [[23,11]])
    print(problem.initial)
    print(problem.domain.head)
    processos = []  # para guardar todos os SearchProcess
    for search_strategy in search_strategies:
        for search_limit in search_limits:
            for search_improve in search_improves:
                for search_heuristic in search_heuristics:
                    sp = SearchProcess(problem, search_strategy, search_limit, search_improve, search_heuristic)
                    sp.launch(max_cpu_time=5.0, monitorization_interval=0.5)  # apenas inicia (NÃO bloqueia)
                    processos.append(sp)



    # Agora que TODOS foram iniciados, eles rodam em paralelo.
    # Vamos esperar cada um terminar, coletar stats e escrever no Excel.
    resultados = []
    for sp in processos:
        sp.wait()  # Aguardar finalização desse SearchProcess
        if sp.stats is not None:
            stats = sp.stats
            stats['elapsed_time'] = sp.search_time
        else:
            # TIMEOUT ou erro
            stats = {
                'Strategy': sp.search_strategy,
                'Improve': sp.search_improve,
                'Limit': sp.search_limit,
                'Heuristic': sp.search_heuristic,
                'Solution': 'TIMEOUT',
                'Terminals': 1,
                'Non-terminals': 'None',
                'Avg Branching': 'None',
                'Cost': 'None',
                'Length': 'None',
                'Num Skipped':  0,
                'Num Solutions': 'None',
                'elapsed_time': sp.search_time,
                'Path': 'None',
                }

        resultados.append(stats)
        # Escrever resultados no Excel
        df = pd.DataFrame(resultados)
        df.to_excel(exel_name, index=False)
        wb = load_workbook(exel_name)
        ws = wb.active
        # Aplicar cor de preenchimento às células onde 'Solution' é 'ERROR: None'
        fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="lightGrid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[4].value == 'FAIL' or row[4].value == 'TIMEOUT':  # Supondo que 'Solution' está na segunda coluna
                for cell in row:
                    cell.fill = fill
        wb.save(exel_name)
    #domain.print_mapa(path = path)
def test_reta_heuristic():
    def rotacionar_ponto(ponto, angulo, ponto_medio):
        # Transladar ponto para a origem
        x, y = ponto
        px, py = ponto_medio
        x_transladado = x - px
        y_transladado = y - py

        # Aplicar matriz de rotação
        angulo_rad = math.radians(angulo)
        x_rotacionado = x_transladado * math.cos(angulo_rad) - y_transladado * math.sin(angulo_rad)
        y_rotacionado = x_transladado * math.sin(angulo_rad) + y_transladado * math.cos(angulo_rad)

        # Transladar ponto de volta
        x_final = x_rotacionado + px
        y_final = y_rotacionado + py

        return (x_final, y_final)
    def rotacionar_reta(ponto1, ponto2, angulo):
        # Calcular ponto médio
        ponto_medio = ((ponto1[0] + ponto2[0]) / 2, (ponto1[1] + ponto2[1]) / 2)

        # Rotacionar pontos
        ponto1_rotacionado = rotacionar_ponto(ponto1, angulo, ponto_medio)
        ponto2_rotacionado = rotacionar_ponto(ponto2, angulo, ponto_medio)

        return ponto1_rotacionado, ponto2_rotacionado
    def calcular_reta(ponto1, ponto2):
        x1, y1 = ponto1
        x2, y2 = ponto2

        # Calcula a inclinação (m)
        if x2 - x1 != 0:
            m = (y2 - y1) / (x2 - x1)
        else:
            raise ValueError("Os pontos têm a mesma coordenada x, a inclinação é indefinida (reta vertical).")

        # Calcula a interseção com o eixo y (b)
        b = y1 - m * x1

        return m, b

    def pontos_inteiros_na_reta(ponto1, ponto2):
        m, b = calcular_reta(ponto1, ponto2)
        print(f" y = {m}x + {b}")
        x1, y1 = ponto1
        x2, y2 = ponto2

        pontos_inteiros = []

        # Iterar sobre os valores de x entre x1 e x2
        for x in range(min(x1, x2), max(x1, x2) + 1):
            y = m * x + b
            ponto = [x, int(y)]
            if ponto not in pontos_inteiros:
                pontos_inteiros.append(ponto)

        # Iterar sobre os valores de y entre y1 e y2
        for y in range(min(y1, y2), max(y1, y2) + 1):
            if m != 0:
                x = (y - b) / m
                ponto = [int(x), y]
                if ponto not in pontos_inteiros:
                    pontos_inteiros.append(ponto)

        pontos_inteiros.sort()
        return pontos_inteiros
    domain = SearchDomain()
    #state = {"size": [10, 10], "map":[[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0]], "fps": 10, "timeout": 3000, "level": 1}
    state = {"size": [20, 20], "map":[[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]], "fps": 10, "timeout": 3000, "level": 1}
    domain.atualize_domain(state)
    # Exemplo de uso
    pontos = []
    head = [5, 5]
    #goal = [20, 20]
    #goal = [20,10]
    goal = [20,5]
    pontos_goal = pontos_inteiros_na_reta(head, goal)
    print(f"Pontos inteiros na reta entre {head} e {goal}: {pontos_goal}")
    goal_rotated1 = rotacionar_ponto(goal, -40, head)
    print("goal_rotated1",goal_rotated1)
    pontos_goal1 = pontos_inteiros_na_reta(head, (int(round(goal_rotated1[0])),int(round(goal_rotated1[1]))))
    goal_rotated2 = rotacionar_ponto(goal, -20, head)
    print("goal_rotated1",goal_rotated2)
    pontos_goal2 = pontos_inteiros_na_reta(head, (int(round(goal_rotated2[0])),int(round(goal_rotated2[1]))))
    goal_rotated3 = rotacionar_ponto(goal, 20, head)
    print("goal_rotated1",goal_rotated3)
    pontos_goal3 = pontos_inteiros_na_reta(head, (int(round(goal_rotated3[0])),int(round(goal_rotated3[1]))))
    goal_rotated4 = rotacionar_ponto(goal, 40, head)
    print("goal_rotated1",goal_rotated4)
    pontos_goal4 = pontos_inteiros_na_reta(head, (int(round(goal_rotated4[0])),int(round(goal_rotated4[1]))))
    pontos = pontos + pontos_goal
    pontos = pontos + pontos_goal1
    pontos = pontos + pontos_goal2
    pontos = pontos + pontos_goal3
    pontos = pontos + pontos_goal4
    domain.print_mapa(path = pontos)
    
if __name__ == "__main__":
    #ToDo: fazer um test_search com pools de processos, um apenas com um monitor para os varios processos (o monitor teria que ter uma prioridade maior que os outros processos)
    #ToDo: fazer um script para salvar dados de mapas, states etc para testar aqui
    """ start_time = monotonic()
    test_search()
    start_end1_time = monotonic() """
    test_search_queue(exel_name="resultados2.xlsx")
    """ end_time = monotonic()
    print(f"test_search finalizado em: {start_end1_time-start_time:.2f} sec")
    print(f"test_search_queue finalizado em: {end_time-start_end1_time:.2f} sec") """
    #test_reta()